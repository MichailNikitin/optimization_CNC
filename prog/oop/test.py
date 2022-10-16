import openpyxl

class Data_Transform:
    def __init__(self, param, row):
        self.param = param
        self.row = row
    def calculation(self):
        print("ввод данных по заготовке и детали")
        dz_max = 50.3  # Максимальный диаметр заготовки, мм
        tdz = 0.039  # допуск диаметра заготовки, мм
        dz_min = dz_max - tdz  # минимальный диаметр заготовки, мм
        dd_max = 50.00  # Максимальный диаметр детали, мм
        rzag_max = dz_max / 2  # максимальный радиус заготовки, мм
        td = 0.016  # допуск диаметра детали, мм
        dd_min = dd_max - td  # минимальный диаметр детали, мм
        pripd_max = dz_max - 0.5 * (dd_max + dd_min)  # припуск на диаметр максимальный, мм
        pripd_min = dz_min - 0.5 * (dd_max + dd_min)  # припуск на диаметр минимальный, мм
        etgb = 9.47  # ei * tg betta см.формулу
        b = 10.0  # длина шлифуемой поверхности мм
        sigma = 180.0  # интенсивность напряжений, кг / мм2
        print("ввод исходных данных по шлифовальному кругу")
        stzat = 0.01  # степень затупления круга
        dk = 300  # диаметр круга, мм
        print(f"dk={dk}   stzat={stzat}")
        print("ввод исходных данных по станку")
        podatl = 0.003  # Податливость технологической системы, мм / кг
        vk = 30000.0 * 60.0  # окружная скорость вращения круга, мм / мин
        nd = 300.0  # частота вращения детали, об / мин
        print(f"nd={nd}  vk={vk}")
        print("ввод исходных данных по циклу")
        time_ob = 1.0 / nd  # время одного оборота заготовки, мин
        zmax = 2  # количество ступеней цикла
        sp.append(3.0)  # программная минутная подача на первой ступени цикла, мм / мин
        sp.append(0.1)  # программная минутная подача на второй ступени цикла, мм / мин
        tp_ob.append(sp[0] / nd)  # программная на оборот первой ступени цикла, мм / об
        tp_ob.append(sp[1] / nd)  # программная на оборот второй ступени цикла, мм / об
        print("ввод исходных данных по ПАКу")
        pripd_pak.append(0.85 * pripd_max)  # припуск на диаметр на первой ступени цикла, мм
        pripd_pak.append(pripd_max - pripd_pak[0])  # припуск на диаметр на второй ступени цикла, мм
        dpak.append(dz_max - pripd_pak[0])  # диаметр ПАК переключения подачи после первой ступени
        dpak.append(0.5 * (dd_max + dd_min))  # диаметр ПАК переключения подачи после второй ступени(конец цикла)
        print(f"nd={nd}  zmax={zmax}  sp[0]={sp[0]}  sp[1]={sp[1]}  tp_ob[0]={tp_ob[0]}  tp_ob[1]={tp_ob[1]}")
        print(f"pripd_pak[0]={pripd_pak[0]}  pripd_pak[1]={pripd_pak[1]}  dpak[0]={dpak[0]}  dpak[1]={dpak[1]}")
        print("расчет коэффициентов")
        k3 = m.pi * dz_max * b * sigma * etgb * nd / vk  # первый коф - т в формуле силы Ру через tф
        k4 = stzat * sigma * b * m.sqrt(dz_max * dk / (dz_max + dk)) / 3.0  # второй коф - т в формуле силы Ру через tф
        с1_tf_ob = 0.5 * podatl * k4 / (1.0 + k3 * podatl)  # коэф в формуле tf_ob
        print(f"k3={k3}   k4={k4}  с1_tf_ob ={с1_tf_ob}")
        print("Закончен ввод данных")
        data_grafik = "datagrafiks2.xlsx"  # имя файла где будут записаны данные для графиков
        f = openpyxl.Workbook()
        f.create_sheet(title='Первый лист', index=0)
        sheet = f["Первый лист"]
        column_names = ["z", "nsum", "sp[z]", "tp_ob[z]", "tp_sum", "tf_ob", "tf_sum", "sf", "py", "ypr_def", "ddo",
                        "dpak[z]", "time_cicle", "tf_sum_i_1"]
        row = 1
        for i in range(1, 15):
            sheet.cell(row=row, column=i, value=column_names[i - 1])
        row += 1
        print("METKA1+++++++++=====Начало расчета цикла==+++++++++++++++")
        print("Задаем стартовые значения переменным")
        ddo = dz_max  # ddo-текущий диаметр детали контролируемый ПАКом  в процессе съема металла
        nsum = 0  # обнуляем номер оборота детали
        time_cicle = 0  # обнуляем время цикла шлифования, мин
        py = 0.0
        ypr_def = 0
        sf = 0.0  # Обнуляем фактическую минутную подачу
        tf_ob = 0.0  # Обнуляем фактическую подачу за оборот
        # одинаково поэтому пока так time_cicle = 0.0  # Обнуляем время цикла
        z = 0  # первая ступень цикла z = 0 имеет индекс 0
        # Обнуляем накопленные подачи
        tp_sum = 0.0  # Обнуляем суммарную программную подачу при nsum = 0, мм
        tf_sum = 0.0  # Обнуляем суммарную фактическую подачу при nsum = 0, мм
        tf_sum_i_1 = 0.0
        sheet.cell(row=row, column=1, value=z);
        sheet.cell(row=row, column=2, value=nsum)
        sheet.cell(row=row, column=3, value=sp[z]);
        sheet.cell(row=row, column=4, value=tp_ob[z])
        sheet.cell(row=row, column=5, value=tp_sum);
        sheet.cell(row=row, column=6, value=tf_ob)
        sheet.cell(row=row, column=7, value=tf_sum);
        sheet.cell(row=row, column=8, value=sf)
        sheet.cell(row=row, column=9, value=py);
        sheet.cell(row=row, column=10, value=ypr_def)
        sheet.cell(row=row, column=11, value=ddo);
        sheet.cell(row=row, column=12, value=dpak[z])
        sheet.cell(row=row, column=13, value=time_cicle);
        sheet.cell(row=row, column=14, value=tf_sum_i_1)
        row += 1
        print("Начинаем расчет цикла шлифования ===========%%%%%%%%%%%%%%%%%%%%%%%%%")
        for z in range(0, zmax):
            while (ddo > dpak[z]):
                nsum = nsum + 1
                tp_sum = tp_sum + tp_ob[z]  # считаем текущее значение накопленной программной подачи на каждом радиусе
                c2_tf_ob = (tp_sum - tf_sum) / (1.0 + k3 * podatl)  # коэф в формуле tf_ob
                x = -с1_tf_ob + m.sqrt(с1_tf_ob * с1_tf_ob + c2_tf_ob)
                tf_ob = x ** 2  # считаем текущее значение фактической подачи
                tf_sum = tf_sum + tf_ob  # считаем для j - го радиуса текущее значение накопленной фактической подачи
                tf_sum_i_1 = tf_sum - tf_ob  # считаем предшествующую накопленную фактическую подачу на i - 1 обороте
                sf = tf_ob * nd  # вычисляем фактическую минутную подачу
                py = k3 * tf_ob + k4 * m.sqrt(tf_ob)  # радиальная сила резания, кг
                ypr_def = podatl * py  # упругая деформация ТС, мм
                ddo = ddo - 2 * tf_ob  # расчет текущего значения диаметра
                time_cicle = nsum * time_ob  # считаем время цикла шлифования, мин
                print(f"METKA 1 z={z}  nsum={nsum}  tp_ob[z]={tp_ob[z]}")
                print(f"tf_ob={tf_ob}  sf={sf}  py={py}  ypr_def={ypr_def}")
                print(f"tp_sum={tp_sum}  tf_sum={tf_sum}  tf_sum_i_1={tf_sum_i_1}")
                print(f"ddo={ddo}  dpak[z]={dpak[z]}  time_cicle={time_cicle}")
                sheet.cell(row=row, column=1, value=z);
                sheet.cell(row=row, column=2, value=nsum)
                sheet.cell(row=row, column=3, value=sp[z]);
                sheet.cell(row=row, column=4, value=tp_ob[z])
                sheet.cell(row=row, column=5, value=tp_sum);
                sheet.cell(row=row, column=6, value=tf_ob)
                sheet.cell(row=row, column=7, value=tf_sum);
                sheet.cell(row=row, column=8, value=sf)
                sheet.cell(row=row, column=9, value=py);
                sheet.cell(row=row, column=10, value=ypr_def)
                sheet.cell(row=row, column=11, value=ddo);
                sheet.cell(row=row, column=12, value=dpak[z])
                sheet.cell(row=row, column=13, value=time_cicle);
                sheet.cell(row=row, column=14, value=tf_sum_i_1)
                row += 1
            print("КОНЕЦ ЦИКЛА while ПО ddo")
        print("КОНЕЦ ЦИКЛА for по z")
        f.save(data_grafik)
        input()

        #тут вычисление новых значений
    def read_param(self):
        f = openpyxl.load_workbook("datagrafiks2.xlsx")
        sheet = f["Первый лист"]
        i = 1
        for key in self.param:
            self.param[key] = sheet.cell(row = self.row, column = i).value()
            i += 1
        f.save("datagrafiks2.xlsx")
    def write_param(self):
        f = openpyxl.load_workbook("datagrafiks2.xlsx")
        sheet = f["Первый лист"]
        i = 1
        for key in self.param:
            sheet.cell(row=self.row, column=i, value=self.param[key])
            i += 1
        f.save("datagrafiks2.xlsx")
    def removing_param(self):
        f = openpyxl.load_workbook("datagrafiks2.xlsx")
        sheet = f["Первый лист"]
        for i in range(1, 15):
            sheet.cell(row=self.row, column=i, value="")
        f.save("datagrafiks2.xlsx")
        self.param.clear()